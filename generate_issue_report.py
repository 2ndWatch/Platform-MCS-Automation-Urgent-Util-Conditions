from string import Template
import requests
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import os


def epoch_time_ms():
    today_dt = datetime.now()
    today = round(datetime.now().timestamp() * 1000)
    # print(f'Today: {today}')

    seven_days_dt = today_dt - timedelta(days=7)
    seven_days = round(seven_days_dt.timestamp() * 1000)
    # print(f'One week ago: {seven_days}')

    return today, seven_days


def get_issue_count(endpoint, headers, client_name, account_id, policy_id, issues_df, logger):
    logger.info('Collecting issue data...')
    today, seven_days = epoch_time_ms()
    seven_days = 1694643300000  # Wed 9/13 16:15 MDT
    saturday = 1694239200000  # Sat 11/9 00:00 MDT
    # one_month = 1691335800000

    issue_template = Template("""
        {
          actor {
            account(id: $account_id) {
              aiIssues {
                issues(
                  timeWindow: {startTime: $start, endTime: $end}
                  cursor: "$cursor"
                ) {
                  issues {
                    conditionName
                    policyIds
                  }
                  nextCursor
                }
              }
            }
          }
        }
        """)

    test_critical_cpu = 0
    test_major_cpu = 0
    test_minor_cpu = 0
    active_major_cpu = 0
    active_minor_cpu = 0
    test_critical_memory = 0
    test_major_memory = 0
    test_minor_memory = 0
    active_major_memory = 0
    active_minor_memory = 0
    total_issue_count = 0

    is_cursor = True
    cursor = ''

    while is_cursor:

        issue_template_fmtd = issue_template.substitute({'account_id': account_id,
                                                         'start': seven_days,
                                                         'end': today,
                                                         'cursor': cursor})
        nr_response = requests.post(endpoint,
                                    headers=headers,
                                    json={'query': issue_template_fmtd}).json()

        # print(nr_response)
        try:
            issues = nr_response['data']['actor']['account']['aiIssues']['issues']['issues']
            total_issue_count += len(issues)
            logger.info(f'   {total_issue_count} issues found. Processing issues...')
            # logger.info(f'      {issues}')

            for issue in issues:
                try:
                    condition_name = issue['conditionName'][0]
                    condition_policy = int(issue['policyIds'][0])
                    # logger.info(f'   Condition {condition_name} in policy {condition_policy}')
                    # logger.info(f'   In test policy: {condition_policy == int(policy_id)}')
                    if 'RDS' not in condition_name and 'rds' not in condition_name \
                            and 'elasticache' not in condition_name:

                        # 2W_CPU_Mem_100 policy
                        if condition_policy == int(policy_id):
                            # logger.info('      Condition is in test policy!')
                            if 'CPU' in condition_name:
                                # logger.info('         Condition is for CPU!')
                                if 'CRITICAL' in condition_name:
                                    # logger.info('            Condition is CRITICAL!')
                                    test_critical_cpu += 1
                                elif 'MAJOR' in condition_name:
                                    # logger.info('            Condition is MAJOR!')
                                    test_major_cpu += 1
                                elif 'MINOR' in condition_name:
                                    # logger.info('            Condition is MINOR!')
                                    test_minor_cpu += 1
                                else:
                                    continue
                            elif 'Memory' in condition_name:
                                if 'CRITICAL' in condition_name:
                                    test_critical_memory += 1
                                elif 'MAJOR' in condition_name:
                                    test_major_memory += 1
                                elif 'MINOR' in condition_name:
                                    test_minor_memory += 1
                                else:
                                    continue

                        # all other policies
                        else:
                            if 'CPU' in condition_name or 'cpu' in condition_name:
                                if 'MINOR' not in condition_name and 'Minor' not in condition_name:
                                    active_major_cpu += 1
                                elif 'MINOR' in condition_name or 'Minor' in condition_name:
                                    active_minor_cpu += 1
                                else:
                                    continue
                            elif 'Memory' in condition_name or 'memory' in condition_name:
                                if 'MINOR' not in condition_name and 'Minor' not in condition_name:
                                    active_major_memory += 1
                                elif 'MINOR' in condition_name or 'Minor' in condition_name:
                                    active_minor_memory += 1
                                else:
                                    continue
                            else:
                                continue
                    else:
                        continue
                except TypeError:
                    continue

            if nr_response['data']['actor']['account']['aiIssues']['issues']['nextCursor']:
                cursor = nr_response['data']['actor']['account']['aiIssues']['issues']['nextCursor']
                # logger.info(f'      Cursor: {cursor}')
            else:
                is_cursor = False

        except TypeError:
            logger.info('   No more issues found.')
            is_cursor = False

    # test_accounts = [3770774, 2672103, 3588235, 2687834, 3498029, 2671646]
    #
    # if account_id not in test_accounts:
    #     test_maj_cpu = '-'
    #     test_min_cpu = '-'
    #     test_maj_mem = '-'
    #     test_min_mem = '-'

    logger.info('   The following issues were found:')
    logger.info(f'      {test_critical_cpu} test critical CPU issues')
    logger.info(f'      {test_major_cpu} test major CPU issues')
    logger.info(f'      {test_minor_cpu} test minor CPU issues')
    logger.info(f'      {active_major_cpu} active major CPU issues')
    logger.info(f'      {active_minor_cpu} active minor CPU issues')
    logger.info(f'      {test_critical_memory} test critical memory issues')
    logger.info(f'      {test_major_memory} test major memory issues')
    logger.info(f'      {test_minor_memory} test minor memory issues')
    logger.info(f'      {active_major_memory} major memory issues')
    logger.info(f'      {active_minor_memory} minor memory issues')
    logger.info('Adding client row to dataframe.')

    # all_critical_cpu = critical_cpu_issues + critical_cpu_filtered

    # 'Client Name', 'Active Minor CPU', 'Active Major CPU', 'Test Minor CPU',
    # 'Test Major CPU', 'Test CPU 100%', 'Active Minor Mem', 'Active Major Mem',
    # 'Test Minor Mem', 'Test Major Mem', 'Test Memory 100%'
    row = [client_name, active_minor_cpu, active_major_cpu, test_minor_cpu, test_major_cpu, test_critical_cpu,
           active_minor_memory, active_major_memory, test_minor_memory, test_major_memory, test_critical_memory]
    issues_df.loc[len(issues_df)] = row

    return issues_df


def generate_report_from_template(issues_df):
    filename = 'CPU Mem Issues Template.xlsx'
    directory_path = os.path.join("output")

    # create a report workbook based on the template
    report_wb = load_workbook(filename)

    issues_sheet = report_wb['Issues']
    issues_sheet._current_row = 1

    # iterate over dataframe and append rows to report workbook
    for i, row in issues_df.iterrows():
        row_list = row.tolist()
        issues_sheet.append(row_list)

    # center-align the data in columns B through I
    for row in issues_sheet[2:issues_sheet.max_row]:  # skip the header
        for i in range(1, 11):
            cell = row[i]
            cell.alignment = Alignment(horizontal='center')

    # shade odd-numbered rows
    for row in issues_sheet[2:37]:  # skip the header
        row_index = row[0].row
        if row_index % 2 > 0:
            for i in range(11):
                cell = row[i]
                # if cols[-1].value:  # need to check the cells have values, otherwise colors the entire row.
                cell.fill = PatternFill(start_color='D9D9D9', fill_type='solid')

    # generate new filename to include current date
    new_filename = filename.replace('Template', f'Weekly {datetime.today().date()}')

    # save the final workbook with the new filename
    report_wb.save(os.path.join(directory_path, new_filename))
    report_wb.close()

    return
